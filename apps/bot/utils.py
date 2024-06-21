"""Модуль disk.py для работы с excel файлами с google drive."""

from io import BytesIO
from typing import Any

import openpyxl
from django.conf import settings
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

from apps.bot.models import Educator

_EXTENDED_TIME_MAPPING = {
    1: '8:00 - 9:30',
    2: '9:40 - 11:10',
    3: '11:30 - 13:00',
    4: '14:10 - 15:40',
    5: '16:00 - 17:30',
    6: '17:40 - 19:10',
}

_FOLDER_ID = '19yyXXullGGMIT3XISiZ33wkDxHJy0zvb'

_SCOPES = ['https://www.googleapis.com/auth/drive']

_TIME_MAPPING = {
    1: '8:00 - 9:30',
    2: '9:40 - 11:10',
    3: '11:30 - 13:00',
    4: '13:10 - 14:40',
    5: '15:00 - 16:30',
    6: '16:40 - 18:10',
    7: '18:20 - 19:50',
}


def get_credentials() -> Credentials:
    """Возвращает учетные данные из файла учетной записи службы с заданными областями."""
    service_account_file = settings.BASE_DIR / 'apps' / 'bot' / 'data' / settings.API_ACCOUNT
    return Credentials.from_service_account_file(
        service_account_file,
        scopes=_SCOPES,
    )


def get_filenames() -> list[dict]:
    """Получает список файлов с Google Drive."""
    drive_service = build('drive', 'v3', credentials=get_credentials())

    results = (
        drive_service.files()
        .list(
            q=f"'{_FOLDER_ID}' in parents",
            fields='nextPageToken, files(id, name)',
        )
        .execute()
    )

    return results.get('files', [])


def download_file(file_id: str, drive_service: Any) -> BytesIO:
    """Скачивает файл с Google Drive."""

    request = drive_service.files().get_media(fileId=file_id)
    return BytesIO(request.execute())


def process_excel(file_content: BytesIO, group_name: str) -> list[tuple[str, str | int, str, bool]]:
    """Обрабатывает содержимое Excel файла, возвращая список данных для заданной группы."""

    wb = openpyxl.load_workbook(file_content)
    results = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):

            class_hour = 'Классный час' in row
            for idx in range(1, len(row), 3):

                if row[idx] == group_name:

                    room_number = row[idx - 1]
                    teacher_name = row[idx + 1]
                    results.append((sheet.title, room_number, teacher_name, class_hour))

    return results


def process_excel2(
    file_content: BytesIO, teacher_name: str,

) -> list[tuple[str, str | int, str, str, bool]]:
    """Обрабатывает содержимое Excel файла, возвращая список данных для заданного преподавателя."""

    wb = openpyxl.load_workbook(file_content)
    results = []
    teacher_last_name = teacher_name.strip().lower()

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):

            class_hour = 'Классный час' in row
            for idx in range(2, len(row), 3):
                if row[idx] and row[idx].strip().lower().startswith(teacher_last_name):

                    room_number = row[idx - 2]
                    group_name = row[idx - 1]
                    full_teacher_name = row[idx]

                    results.append(
                        (sheet.title, room_number, group_name, full_teacher_name, class_hour),
                    )

    return results


def form_schedule(schedule: str) -> str:
    """Формирует текст расписания, добавляя к каждому занятию время его проведения."""

    schedule_text = schedule.split('\n')
    class_hour_day = any('Классный час' in line for line in schedule_text)

    for i, line in enumerate(schedule_text):
        schedule_text_str = ''

        if line.strip() and line[0].isdigit():
            pair_number = int(line[0])

            if class_hour_day and pair_number in _TIME_MAPPING:
                schedule_text_str = (
                    '🕒 ' + schedule_text[i] + f' {_EXTENDED_TIME_MAPPING[pair_number]}'
                )

            elif pair_number in _TIME_MAPPING:
                schedule_text_str = '🕒 ' + schedule_text[i] + f' {_TIME_MAPPING[pair_number]}'
            schedule_text[i] = schedule_text_str

        elif 'Классный час' in line:
            schedule_text[i] = '🕒 ' + line

    return '\n'.join(schedule_text)


def service(name: str, group: str) -> str:
    """Предоставляет расписание для заданной группы."""

    files = get_filenames()

    chosen_file_name = name
    chosen_file = None
    for file in files:
        if file['name'] == chosen_file_name + '.xlsx':
            chosen_file = file
            break

    if chosen_file is None:
        return 'Файл не найден.'

    group_name = group.upper()

    drive_service = build('drive', 'v3', credentials=get_credentials())

    file_content = download_file(chosen_file['id'], drive_service)

    results = process_excel(file_content, group_name)

    if not results:
        return 'Неправильно введен номер группы или занятий нет.'

    message = []
    for sheet_title, room_number_value, teacher_name, _class_hour in results:

        room_number = room_number_value
        if isinstance(room_number, float):

            room_number = int(room_number)
        message.append(
            f'\n{sheet_title},🔑 Кабинет: {room_number},💼 Преподаватель: {teacher_name}\n',
        )
    message2 = f'{group_name.upper()}\n' + ''.join(message).replace(',', '\n')

    return form_schedule(message2)


def search_schedule_by_teacher(name: str, teacher_name: str) -> str:
    """Предоставляет расписание для заданного преподавателя."""

    files = get_filenames()
    chosen_file_name = name
    chosen_file = None

    for file in files:
        if file['name'] == chosen_file_name + '.xlsx':
            chosen_file = file
            break

    if chosen_file is None:
        return 'Файл не найден.'

    drive_service = build('drive', 'v3', credentials=get_credentials())

    file_content = download_file(chosen_file['id'], drive_service)
    results = process_excel2(file_content, teacher_name)

    if not results:
        return 'Неправильно введены данные или занятий нет.'

    message = []
    for sheet_title, room_number_value, group_name, _full_teacher_name, _class_hour in results:

        room_number = room_number_value
        if isinstance(room_number, float):

            room_number = int(room_number)
        message.append(f'\n{sheet_title},🔑 Кабинет: {room_number},💼 Группа: {group_name}\n')
    message2 = f'{teacher_name.capitalize()}\n' + ''.join(message).replace(',', '\n')

    return form_schedule(message2)


def get_fio(value):
    """Предоставляет ФИО для заданной пользователем фамилии преподавателя."""

    try:
        educator = Educator.objects.get(last_name__icontains=value.capitalize())
        return f'{educator.last_name} {educator.first_name} {educator.middle_name}'
    except Educator.DoesNotExist:
        return 'Неверно введена фамилия преподавателя!'

